import openpyxl, csv, ctypes
import matplotlib.pyplot as plt
import numpy as np
import tkinter as tk
import tkinter.filedialog as tkfd
from math import floor, log10


def main(datatype, performance='', SMART='', thermalcoup='', pcp='', chnl=101, volts=3.3):
    def rounding(x):
        return round(x, -int(floor(log10(abs(x)))))

    ctypes.windll.shcore.SetProcessDpiAwareness(1)
    root = tk.Tk()
    root.withdraw()
    root.update()
    save_dest = tkfd.asksaveasfilename(defaultextension=".png", filetypes=[("Portable Network Graphics (.png)", ".png")])
    root.withdraw()
    root.destroy()

    wb = openpyxl.Workbook()
    ws = wb.active
    wb.create_sheet(index=1)
    wb.create_sheet(index=2)
    wb.create_sheet(index=3)
    ws1 = wb[wb.sheetnames[1]]
    ws2 = wb[wb.sheetnames[2]]
    ws3 = wb[wb.sheetnames[3]]

    allSums = []
    stemps = []
    couptemps = []
    pconvals = []
    iops = 0

    if datatype == 'Random':

        print("random data type")
        print('performance file path passed:', performance)

        try:

            # opening the file and converting to xlsx
            with open(performance) as f:
                reader = csv.reader(f, delimiter=',')
                for row in reader:
                    ws.append(row)
            print("finished opening file")

            # row number where read starts
            for i in range(15, ws.max_row):  # iterates from row 15 to max row
                if ws.cell(row=i, column=1).value == "TimeStamp":
                    readStart = i
                    break
            print("found readstart at", readStart)

            # sums of write
            for i in range(15, readStart):
                if ws.cell(row=i, column=4).value != '':  # stop when column D is not empty
                    iops = 0
                    break
                else:
                    iops += float((ws.cell(row=i, column=8)).value)

                    if ws.cell(row=i, column=3).value == 'Worker 16':  # stop every 16 workers
                        allSums.append(round(iops/1000, 1))
                        iops = 0
            print("write datum appended")

            # sums of read
            for i in range(readStart + 1, ws.max_row):
                if ws.cell(row=i, column=4).value != '':  # stop when column D is not empty
                    break
                else:
                    iops += float((ws.cell(row=i, column=8)).value)

                    if ws.cell(row=i, column=3).value == 'Worker 16':  # stop every 16 workers
                        allSums.append(round(iops/1000, 1))
                        iops = 0
            print("read datum appended")

            print("list of all sums:")
            print(allSums)

        except FileNotFoundError:
            print('no random file found')

    elif datatype == 'Sequential':

        print("data type is Sequential")
        print('performance file path passed:', performance)

        try:
            with open(performance) as f:
                reader = csv.reader(f, delimiter=',')
                for row in reader:
                    ws.append(row)
            print("finished opening file")

            # row number where second section starts
            for sct2 in range(15, ws.max_row):  # iterates from row 15 to max row
                if ws.cell(row=sct2, column=1).value == "TimeStamp":
                    break

            section2 = sct2
            print("found readstart at:", section2)

            for write in range(1, ws.max_column+1):
                if ws.cell(row=14, column=write).value == 'Write MBps (Decimal)':
                    break

            for read in range(1, ws.max_column+1):
                if ws.cell(row=14, column=read).value == 'Read MBps (Decimal)':
                    break

            writeMB = write
            readMB = read

            # all sequential vals
            for i in range(15, section2):
                if ws.cell(row=i, column=4).value != '':  # stop when column D is not empty
                    break

                if ws.cell(row=15, column=writeMB).value > ws.cell(row=15, column=readMB).value:
                    allSums.append(float(ws.cell(row=i, column=writeMB).value))
                else:
                    allSums.append(float(ws.cell(row=i, column=readMB).value))

            for i in range(section2 + 1, ws.max_row):
                if ws.cell(row=i, column=4).value != '':  # stop when column D is not empty
                    break

                if ws.cell(row=section2+1, column=writeMB).value > ws.cell(row=section2+1, column=readMB).value:
                    allSums.append(float(ws.cell(row=i, column=writeMB).value))
                else:
                    allSums.append(float(ws.cell(row=i, column=readMB).value))

            print("read datum appended")
            print("list of all sums:")
            print(allSums)
            print('length of allSums:', len(allSums))
        except FileNotFoundError:
            print('no sequential file found')

    try:
        with open(SMART) as t:
            reader = csv.reader(t, delimiter=',')
            for row in reader:
                ws1.append(row)
        print('opened smart file')

        if allSums:
            print('all sums detected')
            for i in range(2, len(allSums) + 2):
                stemps.append(int(ws1.cell(row=i, column=2).value))
            print('appended same number of len of allsums')
            print('values of stemps:')
            print(stemps)

        else:
            print('all sums not detected')
            for i in range(2, ws1.max_row):
                stemps.append(int(ws1.cell(row=i, column=2).value))
            print('appended all the stemps in the file')
            print('values of stemps:')
            print(stemps)

    except FileNotFoundError:
        print('no smart file found')

    try:
        with open(thermalcoup, encoding='utf-16-le') as c:
            reader = csv.reader(c, delimiter=',')
            for row in reader:
                ws2.append(row)
        print("opened thermal coupling file")

        for i in range(1, ws2.max_row + 1):
            for k in range(1, ws2.max_column + 1):
                try:
                    ws2.cell(row=i, column=k).value = float(ws2.cell(row=i, column=k).value)
                except TypeError:
                    continue
                except ValueError:
                    continue

        print('converted all values to float')

        for col in ws2.iter_cols(max_col=1):
            for cell in col:
                if cell.value == 'Scan':
                    scanrow = cell.row
                    print('found scan at row:', scanrow)
                    break

        for pp in range(2, ws2.max_column+1):
            if (str(ws2.cell(row=scanrow, column=pp).value)) == str('Alarm ' + str(chnl)):
                print(True)
                print(scanrow, pp)
                break

        columnstart = pp - 1
        print("will start scanning at column:", columnstart)

        if allSums:
            print('allsums detected')
            print('column start at', columnstart)
            for k in range(scanrow+1, len(allSums)+scanrow+1):
                couptemps.append(float(ws2.cell(row=k, column=columnstart).value))
            print('appended same number as allsums')
        else:
            print('allsums not detected')
            print('column start at', columnstart)
            for j in range(scanrow+1, ws2.max_row+1):
                couptemps.append(float(ws2.cell(row=j, column=columnstart).value))
            print("appended all values in column")

        print(couptemps)
        print(len(couptemps))

    except FileNotFoundError:
        print('thermal coup file not found')

    try:
        with open(pcp) as pc:
            reader = csv.reader(pc, delimiter=',')
            for row in reader:
                ws3.append(row)
        print('opened power consumption file')
        if allSums:
            print('allsums detected')
            for i in range(17, len(allSums) + 17):
                pconvals.append(float(ws3.cell(row=i, column=2).value) * volts)
            print('appending same number of values as allsums')
            print('all values of power consumption:')
            print(pconvals)
        else:
            print('no allSums detected')
            for i in range(17, ws3.max_row):
                pconvals.append(float(ws3.cell(row=i, column=2).value) * volts)
            print('appending all pcon values')
            print('all values of power consump:')
            print(pconvals)

    except FileNotFoundError:
        print('no pcon file detected')

    tp = [allSums, stemps, couptemps, pconvals]
    print(tp)
    tmp = [x for x in tp if x != []]
    print('tp after removal of empty lists:')
    print(tmp)
    print('length of tmp is', len(tmp))

    if not allSums:
        print('allsums is empty therefore time will take the lowest length list')
        time = [3*i for i in range(1, len(min(tmp, key=len))+1)]
        print("time scale:")
        print(time)

        if len(stemps) > len(time):
            while len(stemps) != len(time):
                stemps.pop()

        if len(couptemps) > len(time):
            while len(couptemps) != len(time):
                couptemps.pop()

        if len(pconvals) > len(time):
            while len(pconvals) != len(time):
                pconvals.pop()
    else:
        print('time detected, will have same length as allsums')
        time = [3*i for i in range(1, len(allSums)+1)]
        print('time scale:')
        print(time)

    print(len(allSums), len(stemps), len(couptemps), len(pconvals))

    plt.figure(figsize=(19.2, 10.8))
    fig, ax = plt.subplots()
    ax2 = ax.twinx()
    ax3 = ax.twinx()

    if allSums and (datatype == 'Sequential'):
        ax.set_ylabel('Performance (MB/s)', color='blue')
        ax.plot(time, allSums, color='blue', label='Performance')
        ax.set_yticks(np.arange(0, max(allSums)*1.2, rounding(max(allSums)/7)))
    elif allSums and (datatype == 'Random'):
        ax.set_ylabel('Performance (kIOPS)', color='blue')
        ax.plot(time, allSums, color='blue', label='Performance')
        ax.set_yticks(np.arange(0, max(allSums)*1.2, rounding(max(allSums)/7)))
    elif not allSums:
        ax.set_ylabel('Performance', color='blue')
        ax.set_yticks(np.arange(0, 1100, 100))

    if stemps or couptemps:
        ax2.set_ylabel('Temperature (°C)', color='red')

        try:
            ax2.plot(time, stemps, color='red', label='SMART Temperature')
        except:
            pass

        try:
            ax2.plot(time, couptemps, color='green', label='Thermal Coup. Temp.')
        except:
            pass

        ax2.set_yticks(np.arange(0, 110, 10))
    else:
        ax2.set_ylabel('Temperature (°C)', color='red')
        ax2.set_yticks(np.arange(0, 110, 10))

    if pconvals:
        ax3.set_ylabel('Power Consumption (Watts)', color='orange')
        ax3.plot(time, pconvals, color='orange', label='Power Consumption')
        if volts <= 5:
            ax3.set_yticks(np.arange(0, max(pconvals)*1.2, 0.3))
        elif volts <= 12:
            ax3.set_yticks(np.arange(0, max(pconvals)*1.2, 0.9))

    else:
        ax3.set_yticks(np.arange(0, 10, 1))

    ax3.spines['right'].set_position(('outward', 60))
    ax3.yaxis.set_label_position('right')
    ax3.yaxis.set_ticks_position('right')

    ax.set_xlabel('Time (seconds)', color='black')
    if max(time) > 400:
        ax.set_xticks(np.arange(0, max(time), 30))
    elif max(time) < 400:
        ax.set_xticks(np.arange(0, max(time), 15))

    # chart legend
    handles, labels = [(a + b + c) for a, b, c in zip(ax.get_legend_handles_labels(), ax2.get_legend_handles_labels(), ax3.get_legend_handles_labels())]
    fig.legend(handles, labels, loc="upper right")
    plt.tight_layout()
    plt.savefig(save_dest)
    plt.show()


if __name__ == '__main__':
    main()
