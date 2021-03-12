import csv, openpyxl, tkinter, os
import tkinter.filedialog
import ctypes
from openpyxl.styles import Border, PatternFill, Side, Alignment


def main():
    jobnames = ['Seq-128K-0-1-32', 'Seq-128K-100-1-32', 'Rnd-4K-0-1-1', 'Rnd-4K-0-1-8', 'Rnd-4K-0-1-32',
                'Rnd-4K-0-1-256',
                'Rnd-4K-70-1-32', 'Rnd-4K-100-1-1', 'Rnd-4K-100-1-8', 'Rnd-4K-100-1-32', 'Rnd-4K-100-1-256',
                'Rnd-8K-0-1-32',
                'Rnd-8K-70-1-32', 'Rnd-8K-100-1-32', 'Seq-128K-100-1-256', 'Seq-128K-0-1-256']

    indices = {}

    ctypes.windll.shcore.SetProcessDpiAwareness(1)
    root = tkinter.Tk()
    root.withdraw()
    root.update()
    path = tkinter.filedialog.askopenfilename()
    if path:
        head, tail = os.path.split(path)
        brand = os.path.splitext(tail)[0]
        root.withdraw()
        root.update()
        save_dest = tkinter.filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                         filetypes=[("Excel Spreadsheet (.xlsx)", ".xlsx")])
        root.withdraw()
        root.destroy()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Raw Data'
        wb.create_sheet(index=0, title='Results')
        rp = wb[wb.sheetnames[0]]
        fill = PatternFill(patternType='solid', fgColor='b3daff')
        normal_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                               bottom=Side(style='thin'))

        def setup():
            with open(path) as f:
                read = csv.reader(f, delimiter=',')
                for row in read:
                    ws.append(row)

            for i in range(0, 29):
                for k in range(0, 135):
                    try:
                        ws.cell(row=k, column=i).value = float(ws.cell(row=k, column=i).value)
                    except TypeError:
                        continue
                    except ValueError:
                        continue

            for col in ws.iter_cols():
                for item in col:
                    if item.value in jobnames:
                        indices[item.value] = item.row

            rp.cell(row=2, column=2).value = 'Brand'
            rp.cell(row=14, column=2).value = 'Compressibility %'
            rp.cell(row=2, column=3).value = str(brand)
            rp.cell(row=14, column=3).value = str(brand)
            rp.cell(row=2, column=3).fill = fill
            rp.cell(row=14, column=3).fill = fill

        def seq128():
            rp.cell(row=3, column=2).value = '128K Seq Read QD32 (MB/sec)'
            rp.cell(row=3, column=3).value = round(float(ws.cell(row=indices['Seq-128K-100-1-32'], column=8).value))

            rp.cell(row=4, column=2).value = '128K Seq Write QD32 (MB/sec)'
            rp.cell(row=4, column=3).value = round(float(ws.cell(row=indices['Seq-128K-0-1-32'], column=10).value))

            rp.cell(row=15, column=2).value = '128K SR QD32 (MB/s)'
            rp.cell(row=15, column=3).value = round(float(ws.cell(row=indices['Seq-128K-100-1-32'], column=8).value))

            rp.cell(row=16, column=2).value = '128K SW QD32 (MB/s)'
            rp.cell(row=16, column=3).value = round(float(ws.cell(row=indices['Seq-128K-0-1-32'], column=10).value))

        def rr4kbiops():
            rp.cell(row=5, column=2).value = '4KB Rand Read QD1 (KIOPs)'
            rrqd1 = float(ws.cell(row=indices['Rnd-4K-100-1-1'], column=7).value)
            rp.cell(row=5, column=3).value = round(rrqd1 / 1000, 2)

            rp.cell(row=6, column=2).value = '4KB Rand Read QD8 (KIOPs)'
            rrqd8 = float(ws.cell(row=indices['Rnd-4K-100-1-8'], column=7).value)
            rp.cell(row=6, column=3).value = round(rrqd8 / 1000, 2)

            rp.cell(row=7, column=2).value = '4KB Rand Read QD32 (KIOPs)'
            rrqd32 = float(ws.cell(row=indices['Rnd-4K-100-1-32'], column=7).value)
            rp.cell(row=7, column=3).value = round(rrqd32 / 1000, 2)

            rp.cell(row=17, column=2).value = '4K RR QD32 (KIOPs)'
            rp.cell(row=17, column=3).value = round(rrqd32 / 1000, 2)

            rp.cell(row=18, column=2).value = '4K R70R QD32 (KIOPs)'
            r70rqd32 = float(ws.cell(row=indices['Rnd-4K-70-1-32'], column=7).value) + float(
                ws.cell(row=indices['Rnd-4K-70-1-32'], column=9).value)
            rp.cell(row=18, column=3).value = round(r70rqd32 / 1000, 2)

            rp.cell(row=8, column=2).value = '4KB Rand Read QD256 (KIOPs)'
            rrqd256 = float(ws.cell(row=indices['Rnd-4K-100-1-256'], column=7).value)
            rp.cell(row=8, column=3).value = round(rrqd256 / 1000, 2)

        def rw4kbiops():
            rp.cell(row=9, column=2).value = '4KB Rand Write QD1 (KIOPs)'
            rwqd1 = float(ws.cell(row=indices['Rnd-4K-0-1-1'], column=9).value)
            rp.cell(row=9, column=3).value = round(rwqd1 / 1000, 2)

            rp.cell(row=10, column=2).value = '4KB Rand Write QD8 (KIOPs)'
            rwqd8 = float(ws.cell(row=indices['Rnd-4K-0-1-8'], column=9).value)
            rp.cell(row=10, column=3).value = round(rwqd8 / 1000, 2)

            rp.cell(row=11, column=2).value = '4KB Rand Write QD32 (KIOPs)'
            rwqd32 = float(ws.cell(row=indices['Rnd-4K-0-1-32'], column=9).value)
            rp.cell(row=11, column=3).value = round(rwqd32 / 1000, 2)

            rp.cell(row=12, column=2).value = '4KB Rand Write QD256 (KIOPs)'
            rwqd256 = float(ws.cell(row=indices['Rnd-4K-0-1-256'], column=9).value)
            rp.cell(row=12, column=3).value = round(rwqd256 / 1000, 2)

            rp.cell(row=19, column=2).value = '4K RW QD32 (KIOPs)'
            rp.cell(row=19, column=3).value = round(rwqd32 / 1000, 2)

        def random8kiops():
            rp.cell(row=20, column=2).value = '8KB RR QD32 (KIOPs)'
            r8krrqd32 = float(ws.cell(row=indices['Rnd-8K-100-1-32'], column=7).value)
            rp.cell(row=20, column=3).value = round(r8krrqd32 / 1000, 2)

            rp.cell(row=21, column=2).value = '8K R70R QD32 (KIOPs)'
            r8k70 = float(ws.cell(row=indices['Rnd-8K-70-1-32'], column=7).value) + float(
                ws.cell(row=indices['Rnd-8K-70-1-32'], column=9).value)
            rp.cell(row=21, column=3).value = round(r8k70 / 1000, 2)

            rp.cell(row=22, column=2).value = '8KB RW QD32 (KIOPs)'
            r8krwqd32 = float(ws.cell(row=indices['Rnd-8K-0-1-32'], column=9).value)
            rp.cell(row=22, column=3).value = round(r8krwqd32 / 1000, 2)

        def latency4k():
            rp.cell(row=23, column=2).value = '4K RR Latency QD1 (usec)'
            lat4krrqd1 = float(ws.cell(row=indices['Rnd-4K-100-1-1'], column=11).value)
            rp.cell(row=23, column=3).value = round(lat4krrqd1)

            rp.cell(row=24, column=2).value = '4K RW Latency QD1 (usec)'
            lat4krwqd1 = float(ws.cell(row=indices['Rnd-4K-0-1-1'], column=20).value)
            rp.cell(row=24, column=3).value = round(lat4krwqd1)

        def latency4k99():
            rp.cell(row=25, column=2).value = '4K RR QD32 99% CI (usec)'
            lat4krrqd32_99 = float(ws.cell(row=indices['Rnd-4K-100-1-32'], column=14).value)
            rp.cell(row=25, column=3).value = round(lat4krrqd32_99)

            rp.cell(row=26, column=2).value = '4K R70R QD32 99% CI (usec)'
            lat4kr70rqd32_99 = float(ws.cell(row=indices['Rnd-4K-70-1-32'], column=14).value) + float(
                ws.cell(row=indices['Rnd-4K-70-1-32'], column=23).value)
            rp.cell(row=26, column=3).value = round(lat4kr70rqd32_99)

            rp.cell(row=27, column=2).value = '4K RW QD32 99% CI (usec)'
            lat4krwqd32_99 = float(ws.cell(row=indices['Rnd-4K-0-1-32'], column=23).value)
            rp.cell(row=27, column=3).value = round(lat4krwqd32_99)

        def latency4k9999():
            rp.cell(row=28, column=2).value = '4K RR QD32 99.99% CI (usec)'
            lat4krrqd32_9999 = float(ws.cell(row=indices['Rnd-4K-100-1-32'], column=16).value)
            rp.cell(row=28, column=3).value = round(lat4krrqd32_9999)

            rp.cell(row=29, column=2).value = '4K R70R QD32 99.99% CI (usec)'
            lat4kr70rqd32_9999 = float(ws.cell(row=indices['Rnd-4K-70-1-32'], column=16).value) + float(
                ws.cell(row=indices['Rnd-4K-70-1-32'], column=25).value)
            rp.cell(row=29, column=3).value = round(lat4kr70rqd32_9999)

            rp.cell(row=30, column=2).value = '4K RW QD32 99.99% CI (usec)'
            lat4krwqd32_9999 = float(ws.cell(row=indices['Rnd-4K-0-1-32'], column=25).value)
            rp.cell(row=30, column=3).value = round(lat4krwqd32_9999)

        def seq128qd256():
            try:
                srqd256 = float(ws.cell(row=indices['Seq-128K-100-1-256'], column=8).value)
                rp.cell(row=31, column=2).value = '128K SR QD256 (MB/S)'
                rp.cell(row=31, column=3).value = round(srqd256)

                for i in range(31, 33):
                    rp.cell(row=i, column=2).border = normal_border
                    rp.cell(row=i, column=3).border = normal_border
                    rp.cell(row=i, column=2).alignment = Alignment(horizontal='center')
                    rp.cell(row=i, column=3).alignment = Alignment(horizontal='center')

            except KeyError:
                pass

            try:
                swqd256 = float(ws.cell(row=indices['Seq-128K-0-1-256'], column=10).value)
                rp.cell(row=32, column=2).value = '128K SW QD256 (MB/S)'
                rp.cell(row=32, column=3).value = round(swqd256)

                for i in range(31, 33):
                    rp.cell(row=i, column=2).border = normal_border
                    rp.cell(row=i, column=3).border = normal_border
                    rp.cell(row=i, column=2).alignment = Alignment(horizontal='center')
                    rp.cell(row=i, column=3).alignment = Alignment(horizontal='center')

            except KeyError:
                pass

        def bordering_align():
            for i in range(2, 31):
                if i == 13:
                    continue
                rp.cell(row=i, column=2).border = normal_border
                rp.cell(row=i, column=3).border = normal_border
                rp.cell(row=i, column=2).alignment = Alignment(horizontal='center')
                rp.cell(row=i, column=3).alignment = Alignment(horizontal='center')
                rp.column_dimensions['B'].width = 30
                rp.column_dimensions['C'].width = 15

        setup()
        bordering_align()
        seq128()
        rr4kbiops()
        rw4kbiops()
        random8kiops()
        latency4k()
        latency4k99()
        latency4k9999()
        seq128qd256()
        wb.save(save_dest)
    else:
        root.destroy()


if __name__ == '__main__':
    main()
