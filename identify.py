from tkinter import *
import ctypes
import tkinter.filedialog
from openpyxl import Workbook
from openpyxl.styles import PatternFill

def main():
    try:
        ctypes.windll.shcore.SetProcessDpiAwareness(1)
        root = Tk()
        root.withdraw()
        root.update()
        paths = tkinter.filedialog.askopenfilenames()
        if paths:
            root.withdraw()
            root.update()
            save_dest = tkinter.filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Spreadsheet (.xlsx)", ".xlsx")])
            root.withdraw()
            root.destroy()

            wb = Workbook()
            ws = wb.active
            ws.title = "ATA Identify Data"
            fill = PatternFill(patternType='solid', fgColor='b3daff')

            d = [[] for x in paths]
            e = []

            cnt = 1
            count = 1
            with open(str(paths[0]), 'r', encoding='utf-8') as f:
                lines = f.readlines()
                for line in lines:
                    d[0].append(line)

            for i in range(1, len(paths)):
                with open(str(paths[i]), 'r', encoding=' utf-8') as g:
                    lines = g.readlines()
                    for line in lines:
                        d[i].append(line)

            for i in range(0, len(paths) - 1):
                a = set(d[i]).symmetric_difference(d[i + 1])
                for x in sorted(a):
                    e.append(x)

            for path in range(0, len(paths)):
                for row in range(0, len(d[path])):
                    r = ws.cell(row=cnt, column=count, value=(d[path])[row])
                    if (d[path])[row] in e:
                        r.fill = fill
                    cnt += 1
                count += 1
                cnt = 1

            wb.save(save_dest)
        else:
            root.destroy()

    except IndexError:
        print("No file detected. Please try again.")

    finally:
        pass


if __name__ == '__main__':
    main()
